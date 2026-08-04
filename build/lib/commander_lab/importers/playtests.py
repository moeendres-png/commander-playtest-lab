from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook

from commander_lab.models import PlaytestParticipant, RealPlaytest


class RealPlaytestImporter:
    """Import one-row-per-player playtest sheets from CSV or XLSX."""

    REQUIRED_COLUMNS = {"game_id", "player_id", "deck_name", "seat"}

    def import_file(self, path: str | Path, *, sheet_name: str | None = None) -> list[RealPlaytest]:
        path_obj = Path(path)
        if path_obj.suffix.casefold() == ".csv":
            with path_obj.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
        elif path_obj.suffix.casefold() in {".xlsx", ".xlsm"}:
            workbook = load_workbook(path_obj, read_only=True, data_only=True)
            try:
                worksheet = workbook[sheet_name] if sheet_name else workbook.active
                values = list(worksheet.iter_rows(values_only=True))
                if not values:
                    return []
                headers = [str(value or "").strip() for value in values[0]]
                rows = [
                    {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                    for row in values[1:]
                    if any(value is not None and str(value).strip() for value in row)
                ]
            finally:
                workbook.close()
        else:
            raise ValueError(f"unsupported playtest format: {path_obj.suffix}")
        return self.import_rows(rows, source_file=str(path_obj))

    def import_rows(
        self,
        rows: Iterable[Mapping[str, object]],
        *,
        source_file: str | None = None,
    ) -> list[RealPlaytest]:
        rows = list(rows)
        if not rows:
            return []
        missing = self.REQUIRED_COLUMNS - set(rows[0].keys())
        if missing:
            raise ValueError(f"playtest rows missing required columns: {sorted(missing)}")

        grouped: dict[str, list[Mapping[str, object]]] = defaultdict(list)
        for row in rows:
            grouped[str(row["game_id"])].append(row)

        games: list[RealPlaytest] = []
        for game_id, game_rows in sorted(grouped.items()):
            participants: list[PlaytestParticipant] = []
            for row in game_rows:
                commander_names = [
                    name.strip()
                    for name in str(row.get("commander_names") or "").split("+")
                    if name.strip()
                ]
                participants.append(
                    PlaytestParticipant(
                        player_id=str(row["player_id"]),
                        player_name=str(row.get("player_name") or "") or None,
                        deck_name=str(row["deck_name"]),
                        commander_names=commander_names,
                        seat=int(row["seat"]),
                        placement=int(row["placement"]) if row.get("placement") not in {None, ""} else None,
                        final_life=int(row["final_life"]) if row.get("final_life") not in {None, ""} else None,
                        mulligans=int(row["mulligans"]) if row.get("mulligans") not in {None, ""} else None,
                        notes=str(row.get("player_notes") or "") or None,
                    )
                )
            first = game_rows[0]
            winners = [
                participant.player_id
                for participant in participants
                if participant.placement == 1
            ]
            played_on = (
                date.fromisoformat(str(first["played_on"]))
                if first.get("played_on") not in {None, ""}
                else None
            )
            games.append(
                RealPlaytest(
                    game_id=game_id,
                    played_on=played_on,
                    pod_size=len(participants),
                    participants=participants,
                    turns=int(first["turns"]) if first.get("turns") not in {None, ""} else None,
                    winner_player_ids=winners,
                    end_reason=str(first.get("end_reason") or "") or None,
                    starting_player_id=str(first.get("starting_player_id") or "") or None,
                    freeform_log=str(first.get("freeform_log") or "") or None,
                    source_file=source_file,
                    validated=False,
                )
            )
        return games
