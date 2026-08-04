from __future__ import annotations

from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook

from commander_lab.models import Deck

from .base import CatalogAwareImporter, DeckImportOptions, ImportErrorWithContext
from .csv_importer import CsvDeckImporter
from .xlsx_importer import XlsxDeckImporter


class GoogleDriveExportImporter(CatalogAwareImporter):
    """Read-only importer for an XLSX export of the current Google Drive workbook.

    The importer never uses the Drive API and never writes to the workbook. It detects deck sheets
    by explicit mapping first, then by common current-project sheet names.
    """

    DEFAULT_SHEET_MAP = {
        "korvold/current": ("01_Korvold", "Korvold", "Korvold_final_100"),
        "rogshai/current": ("02_RogShai", "RogShai", "Ishai_Rograkh_final_100"),
    }

    def import_decks(
        self,
        path: str | Path,
        options_by_deck_id: Mapping[str, DeckImportOptions],
        *,
        sheet_map: Mapping[str, str] | None = None,
    ) -> dict[str, Deck]:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            result: dict[str, Deck] = {}
            for deck_id, options in options_by_deck_id.items():
                sheet_name = self._select_sheet(
                    workbook.sheetnames,
                    deck_id,
                    explicit=sheet_map.get(deck_id) if sheet_map else None,
                )
                rows = XlsxDeckImporter._worksheet_rows(
                    workbook[sheet_name], source_path=f"{path}#{sheet_name}"
                )
                result[deck_id] = CsvDeckImporter(self.catalog).import_rows(
                    rows,
                    options,
                    source_path=f"{path}#{sheet_name}",
                )
            return result
        finally:
            workbook.close()

    @classmethod
    def _select_sheet(
        cls, available: list[str], deck_id: str, *, explicit: str | None = None
    ) -> str:
        if explicit:
            if explicit not in available:
                raise ImportErrorWithContext(
                    f"requested sheet {explicit!r} not found; available={available}"
                )
            return explicit
        candidates = cls.DEFAULT_SHEET_MAP.get(deck_id, ())
        for candidate in candidates:
            if candidate in available:
                return candidate
        folded = {name.casefold(): name for name in available}
        key_fragment = deck_id.split("/")[0].casefold()
        for folded_name, original in folded.items():
            if key_fragment in folded_name:
                return original
        raise ImportErrorWithContext(f"no sheet found for {deck_id}; available={available}")
