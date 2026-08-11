from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from commander_lab.importers import (
    CsvDeckImporter,
    DeckImportOptions,
    GoogleDriveExportImporter,
    ImportErrorWithContext,
    OpponentProfileImporter,
    XlsxDeckImporter,
)


def _small_valid_rows() -> list[dict[str, object]]:
    return [
        {"card_name": "Korvold, Fae-Cursed King", "quantity": 1, "zone": "commander"},
        {"card_name": "Forest", "quantity": 99, "zone": "main"},
    ]


def test_csv_importer(tmp_path: Path, catalog) -> None:
    path = tmp_path / "deck.csv"
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["card_name", "quantity", "zone"])
        writer.writeheader()
        writer.writerows(_small_valid_rows())
    deck = CsvDeckImporter(catalog).import_file(
        path,
        DeckImportOptions(
            deck_id="test/csv",
            name="CSV test",
            commander_names=("Korvold, Fae-Cursed King",),
        ),
    )
    assert deck.total_cards == 100


def test_xlsx_importer_detects_header_after_title(tmp_path: Path, catalog) -> None:
    path = tmp_path / "deck.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Deck"
    sheet.append(["Deck export"])
    sheet.append([])
    sheet.append(["card_name", "quantity", "zone"])
    for row in _small_valid_rows():
        sheet.append([row["card_name"], row["quantity"], row["zone"]])
    workbook.save(path)
    deck = XlsxDeckImporter(catalog).import_file(
        path,
        DeckImportOptions(
            deck_id="test/xlsx",
            name="XLSX test",
            commander_names=("Korvold, Fae-Cursed King",),
        ),
    )
    assert deck.library_cards == 99


def test_google_drive_export_importer_is_read_only_and_maps_sheets(tmp_path: Path, catalog) -> None:
    path = tmp_path / "drive-export.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "01_Korvold"
    second = workbook.create_sheet("02_RogShai")
    for sheet, rows in [
        (first, _small_valid_rows()),
        (
            second,
            [
                {"card_name": "Ishai, Ojutai Dragonspeaker", "quantity": 1, "zone": "commander"},
                {"card_name": "Rograkh, Son of Rohgahh", "quantity": 1, "zone": "commander"},
                {"card_name": "Island", "quantity": 98, "zone": "main"},
            ],
        ),
    ]:
        sheet.append(["card_name", "quantity", "zone"])
        for row in rows:
            sheet.append([row["card_name"], row["quantity"], row["zone"]])
    workbook.save(path)
    before = path.read_bytes()

    importer = GoogleDriveExportImporter(catalog)
    decks = importer.import_decks(
        path,
        {
            "rogshai/current": DeckImportOptions(
                deck_id="rogshai/current",
                name="RogShai",
                commander_names=("Ishai, Ojutai Dragonspeaker", "Rograkh, Son of Rohgahh"),
                uses_partner=True,
            ),
        },
    )
    assert decks["rogshai/current"].total_cards == 100
    with pytest.raises(ImportErrorWithContext, match="retired own deck"):
        importer.import_decks(
            path,
            {
                "korvold/current": DeckImportOptions(
                    deck_id="korvold/current",
                    name="Retired Korvold",
                    commander_names=("Korvold, Fae-Cursed King",),
                )
            },
        )
    assert path.read_bytes() == before


def test_opponent_profile_importer(tmp_path: Path) -> None:
    path = tmp_path / "profile.json"
    path.write_text(
        json.dumps(
            {
                "profile_id": "cosmic/mid-budget-variant-a",
                "name": "Cosmic Spider-Man synthetic completion",
                "commander": {"commanders": ["Cosmic Spider-Man"], "uses_partner": False},
                "list_status": "synthetic_completion",
                "known_cards": [],
                "uncertainty": {
                    "confidence": 0.4,
                    "known_card_count": 20,
                    "synthetic_card_count": 80,
                    "assumptions": ["Unknown slots are mid-budget synthetic assumptions."],
                },
                "data_quality": "synthetic_assumption",
            }
        ),
        encoding="utf-8",
    )
    profiles = OpponentProfileImporter().import_file(path)
    assert profiles[0].uncertainty.synthetic_card_count == 80
