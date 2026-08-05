from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import load_workbook

from commander_lab.models import PlaytestParticipant, RealPlaytest


_COLUMN_ALIASES = {
    "game": "game_id",
    "spiel": "game_id",
    "spiel_id": "game_id",
    "datum": "played_on",
    "podgrosse": "pod_size",
    "podgroesse": "pod_size",
    "datenversion": "dataset_version",
    "player": "player_id",
    "spieler": "player_id",
    "spieler_id": "player_id",
    "name": "player_name",
    "spieler_name": "player_name",
    "deck": "deck_name",
    "version": "deck_version",
    "deckversion": "deck_version",
    "deck_hash": "deck_hash",
    "commander": "commander_names",
    "kommandeur": "commander_names",
    "sitz": "seat",
    "platz": "placement",
    "platzierung": "placement",
    "leben": "final_life",
    "mulligan": "mulligans",
    "starthandlaender": "starting_hand_lands",
    "starthandlander": "starting_hand_lands",
    "starthand_lander": "starting_hand_lands",
    "starthandländer": "starting_hand_lands",
    "laender": "lands_played",
    "länder": "lands_played",
    "erster_rampzug": "first_ramp_turn",
    "ramp": "ramp_events",
    "erster_commander_cast": "first_commander_cast_turn",
    "commander_cast": "first_commander_cast_turn",
    "commander_casts": "commander_casts",
    "commander_entfernungen": "commander_removals_received",
    "removal": "removal_events",
    "erster_drawmotor": "first_independent_draw_engine_turn",
    "drawengines": "independent_draw_engines",
    "draw_engine": "independent_draw_engines",
    "wipes": "boardwipes_cast",
    "boardwipes": "boardwipes_cast",
    "wipes_gesehen": "boardwipes_seen",
    "rebuild": "successful_rebuilds",
    "rebuild_erfolgreich": "rebuilt_after_wipe",
    "ishai_wachstum": "ishai_peak_power",
    "ishai_power": "ishai_peak_power",
    "ishai_power_by_turn": "ishai_power_by_turn",
    "korvold_draws": "korvold_cards_drawn",
    "archenemy": "was_archenemy",
    "archenemy_ereignisse": "archenemy_events",
    "siegachse": "win_axis",
    "niederlagenursache": "loss_causes",
    "niederlagenursachen": "loss_causes",
    "tote_karten": "dead_cards",
    "sequencingfehler": "sequencing_errors",
    "startspieler": "starting_player_id",
    "spielzuege": "turns",
    "spielzuge": "turns",
    "spielzüge": "turns",
    "sieggrund": "end_reason",
    "notizen": "player_notes",
}


class RealPlaytestImporter:
    """Import versioned, one-row-per-player real playtest sheets.

    CSV and XLSX are read-only. German and English column aliases are normalized. Missing
    calibration fields do not fabricate values; the game is imported with explicit validation
    errors and is excluded from calibration until corrected.
    """

    REQUIRED_COLUMNS = {"game_id", "player_id", "deck_name", "seat"}

    def import_file(
        self,
        path: str | Path,
        *,
        sheet_name: str | None = None,
        dataset_version: str | None = None,
    ) -> list[RealPlaytest]:
        path_obj = Path(path).resolve()
        suffix = path_obj.suffix.casefold()
        if suffix == ".csv":
            with path_obj.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
        elif suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(path_obj, read_only=True, data_only=True)
            try:
                worksheet = workbook[sheet_name] if sheet_name else workbook.active
                values = list(worksheet.iter_rows(values_only=True))
                if not values:
                    return []
                headers = [str(value or "").strip() for value in values[0]]
                rows = [
                    {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                    for row in values[1:]
                    if any(value is not None and str(value).strip() for value in row)
                ]
            finally:
                workbook.close()
        elif suffix == ".json":
            payload = json.loads(path_obj.read_text(encoding="utf-8"))
            if isinstance(payload, dict) and "games" in payload:
                payload = payload["games"]
            if not isinstance(payload, list):
                raise ValueError("playtest JSON must contain a list of row objects or games")
            if payload and isinstance(payload[0], dict) and "participants" in payload[0]:
                games = [RealPlaytest.model_validate(item) for item in payload]
                for game in games:
                    game.source_file = str(path_obj)
                    game.source_sha256 = self._file_sha256(path_obj)
                return games
            rows = payload
        else:
            raise ValueError(f"unsupported playtest format: {path_obj.suffix}")
        return self.import_rows(
            rows,
            source_file=str(path_obj),
            source_sha256=self._file_sha256(path_obj),
            dataset_version=dataset_version,
        )

    def import_rows(
        self,
        rows: Iterable[Mapping[str, object]],
        *,
        source_file: str | None = None,
        source_sha256: str | None = None,
        dataset_version: str | None = None,
    ) -> list[RealPlaytest]:
        normalized_rows = [self._normalize_row(row) for row in rows]
        if not normalized_rows:
            return []
        missing = self.REQUIRED_COLUMNS - set(normalized_rows[0])
        if missing:
            raise ValueError(f"playtest rows missing required columns: {sorted(missing)}")

        grouped: dict[str, list[Mapping[str, object]]] = defaultdict(list)
        for row in normalized_rows:
            game_id = self._text(row.get("game_id"))
            if not game_id:
                raise ValueError("game_id cannot be empty")
            grouped[game_id].append(row)

        games: list[RealPlaytest] = []
        for game_id, game_rows in sorted(grouped.items()):
            participants = [self._participant(row) for row in game_rows]
            first = game_rows[0]
            winners = [participant.player_id for participant in participants if participant.placement == 1]
            played_on = self._date(first.get("played_on"))
            supplied_pod_size = self._int(first.get("pod_size"))
            resolved_dataset_version = (
                dataset_version
                or self._text(first.get("dataset_version"))
                or self._text(first.get("data_version"))
                or "unversioned"
            )
            resolved_starting_player = self._starting_player(game_rows, participants)
            validation_errors = self._validation_errors(
                participants,
                turns=self._int(first.get("turns")),
                starting_player_id=resolved_starting_player,
                dataset_version=resolved_dataset_version,
            )
            game = RealPlaytest(
                dataset_version=resolved_dataset_version,
                game_id=game_id,
                played_on=played_on,
                pod_size=supplied_pod_size or len(participants),
                participants=participants,
                turns=self._int(first.get("turns")),
                winner_player_ids=winners,
                end_reason=self._text(first.get("end_reason")),
                starting_player_id=resolved_starting_player,
                freeform_log=self._text(first.get("freeform_log")),
                source_file=source_file,
                source_sha256=source_sha256,
                validated=not validation_errors,
                validation_errors=validation_errors,
            )
            games.append(game)
        return games

    def _participant(self, row: Mapping[str, object]) -> PlaytestParticipant:
        commander_names = self._list(row.get("commander_names"), separators=("+", "|", ";"))
        placement = self._int(row.get("placement"))
        return PlaytestParticipant(
            player_id=self._required_text(row, "player_id"),
            player_name=self._text(row.get("player_name")),
            deck_name=self._required_text(row, "deck_name"),
            deck_version=self._text(row.get("deck_version")) or "unversioned",
            deck_hash=self._text(row.get("deck_hash")),
            commander_names=commander_names,
            seat=self._required_int(row, "seat"),
            placement=placement,
            final_life=self._int(row.get("final_life")),
            mulligans=self._int(row.get("mulligans")),
            starting_hand_lands=self._int(row.get("starting_hand_lands")),
            lands_played=self._int(row.get("lands_played")),
            first_ramp_turn=self._int(row.get("first_ramp_turn")),
            ramp_events=self._int(row.get("ramp_events")),
            first_commander_cast_turn=self._int(row.get("first_commander_cast_turn")),
            commander_casts=self._int(row.get("commander_casts")),
            commander_removals_received=self._int(row.get("commander_removals_received")),
            removal_events=self._int(row.get("removal_events")),
            first_independent_draw_engine_turn=self._int(
                row.get("first_independent_draw_engine_turn")
            ),
            independent_draw_engines=self._int(row.get("independent_draw_engines")),
            boardwipes_cast=self._int(row.get("boardwipes_cast")),
            boardwipes_seen=self._int(row.get("boardwipes_seen")),
            successful_rebuilds=self._int(row.get("successful_rebuilds")),
            rebuilt_after_wipe=self._bool(row.get("rebuilt_after_wipe")),
            ishai_peak_power=self._float(row.get("ishai_peak_power")),
            ishai_power_by_turn=self._turn_series(row.get("ishai_power_by_turn")),
            korvold_cards_drawn=self._int(row.get("korvold_cards_drawn")),
            was_archenemy=self._bool(row.get("was_archenemy")),
            archenemy_events=self._int(row.get("archenemy_events")),
            win_axis=self._text(row.get("win_axis")),
            loss_causes=[] if placement == 1 else self._list(row.get("loss_causes")),
            dead_cards=self._list(row.get("dead_cards")),
            sequencing_errors=self._list(row.get("sequencing_errors")),
            notes=self._text(row.get("player_notes") or row.get("notes")),
        )

    @staticmethod
    def _validation_errors(
        participants: list[PlaytestParticipant],
        *,
        turns: int | None,
        starting_player_id: str | None,
        dataset_version: str,
    ) -> list[str]:
        errors: list[str] = []
        if dataset_version == "unversioned":
            errors.append("dataset_version_missing")
        if turns is None:
            errors.append("turn_count_missing")
        if not starting_player_id:
            errors.append("starting_player_missing")
        for participant in participants:
            prefix = participant.player_id
            if participant.deck_version == "unversioned":
                errors.append(f"{prefix}:deck_version_missing")
            if participant.placement is None:
                errors.append(f"{prefix}:placement_missing")
            if participant.mulligans is None:
                errors.append(f"{prefix}:mulligans_missing")
            if participant.starting_hand_lands is None:
                errors.append(f"{prefix}:starting_hand_lands_missing")
        return sorted(set(errors))

    @classmethod
    def _normalize_row(cls, row: Mapping[str, object]) -> dict[str, object]:
        normalized: dict[str, object] = {}
        for key, value in row.items():
            canonical = cls._column_key(str(key))
            normalized[_COLUMN_ALIASES.get(canonical, canonical)] = value
        return normalized

    @staticmethod
    def _column_key(value: str) -> str:
        value = value.strip().casefold().replace("-", "_").replace(" ", "_")
        value = value.replace("ä", "a").replace("ö", "o").replace("ü", "u").replace("ß", "ss")
        return re.sub(r"_+", "_", value).strip("_")

    @staticmethod
    def _file_sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _text(value: object) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @classmethod
    def _required_text(cls, row: Mapping[str, object], key: str) -> str:
        value = cls._text(row.get(key))
        if value is None:
            raise ValueError(f"{key} cannot be empty")
        return value

    @staticmethod
    def _int(value: object) -> int | None:
        if value is None or str(value).strip() == "":
            return None
        return int(float(str(value).strip()))

    @classmethod
    def _required_int(cls, row: Mapping[str, object], key: str) -> int:
        value = cls._int(row.get(key))
        if value is None:
            raise ValueError(f"{key} cannot be empty")
        return value

    @staticmethod
    def _float(value: object) -> float | None:
        if value is None or str(value).strip() == "":
            return None
        return float(str(value).strip().replace(",", "."))

    @staticmethod
    def _bool(value: object) -> bool | None:
        if value is None or str(value).strip() == "":
            return None
        normalized = str(value).strip().casefold()
        if normalized in {"1", "true", "yes", "ja", "y"}:
            return True
        if normalized in {"0", "false", "no", "nein", "n"}:
            return False
        raise ValueError(f"invalid boolean value: {value!r}")

    @staticmethod
    def _date(value: object) -> date | None:
        if value is None or str(value).strip() == "":
            return None
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value).strip()[:10])

    @classmethod
    def _list(
        cls,
        value: object,
        *,
        separators: tuple[str, ...] = ("|", ";"),
    ) -> list[str]:
        if value is None or str(value).strip() == "":
            return []
        if isinstance(value, (list, tuple)):
            return [str(item).strip() for item in value if str(item).strip()]
        text = str(value).strip()
        pattern = "|".join(re.escape(separator) for separator in separators)
        return [item.strip() for item in re.split(pattern, text) if item.strip()]

    @classmethod
    def _turn_series(cls, value: object) -> dict[int, float]:
        if value is None or str(value).strip() == "":
            return {}
        if isinstance(value, dict):
            return {int(turn): float(power) for turn, power in value.items()}
        text = str(value).strip()
        try:
            payload = json.loads(text)
        except json.JSONDecodeError:
            payload = {}
            for item in cls._list(text):
                turn, power = item.split(":", 1)
                payload[int(turn.strip())] = float(power.strip().replace(",", "."))
        if not isinstance(payload, dict):
            raise ValueError("ishai_power_by_turn must be a JSON object or turn:power list")
        return {int(turn): float(power) for turn, power in payload.items()}

    @classmethod
    def _starting_player(
        cls,
        rows: list[Mapping[str, object]],
        participants: list[PlaytestParticipant],
    ) -> str | None:
        first = rows[0]
        explicit = cls._text(first.get("starting_player_id"))
        if explicit:
            return explicit
        for participant, row in zip(participants, rows, strict=True):
            if cls._bool(row.get("is_starting_player")):
                return participant.player_id
        return None
