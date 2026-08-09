from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from commander_lab.models import Deck

from .base import CatalogAwareImporter, DeckImportOptions, ImportErrorWithContext
from .csv_importer import CsvDeckImporter


class XlsxDeckImporter(CatalogAwareImporter):
    def import_file(
        self,
        path: str | Path,
        options: DeckImportOptions,
        *,
        sheet_name: str | None = None,
    ) -> Deck:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            if not isinstance(worksheet, Worksheet):
                raise ImportErrorWithContext("selected sheet is not a worksheet", source=path)
            rows = self._worksheet_rows(worksheet, source_path=path)
            return CsvDeckImporter(self.catalog).import_rows(rows, options, source_path=path)
        finally:
            workbook.close()

    @staticmethod
    def _worksheet_rows(
        worksheet: Worksheet,
        *,
        source_path: str | Path | None = None,
    ) -> list[dict[str, object]]:
        materialized = [list(row) for row in worksheet.iter_rows(values_only=True)]
        if not materialized:
            raise ImportErrorWithContext("worksheet is empty", source=source_path)

        header_index = None
        headers: list[str] = []
        accepted = {"oracle_name", "card_name", "card", "name", "karte", "kartenname"}
        for index, row in enumerate(materialized[:25]):
            candidate = [str(value).strip() if value is not None else "" for value in row]
            if any(value.casefold() in accepted for value in candidate):
                header_index = index
                headers = candidate
                break
        if header_index is None:
            raise ImportErrorWithContext(
                "could not detect a header row containing a card-name column",
                source=source_path,
            )

        result: list[dict[str, object]] = []
        for row in materialized[header_index + 1 :]:
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            record: dict[str, object] = {
                header: row[column] if column < len(row) else None
                for column, header in enumerate(headers)
                if header
            }
            result.append(record)
        return result
